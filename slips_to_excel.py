import logging
import os
import re
from typing import Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from pymupdf4llm import to_markdown
from rich.console import Console
from rich.logging import RichHandler

console = Console()
logging.basicConfig(
    level=logging.INFO,
    format="%(message)s",
    handlers=[
        RichHandler(console=console, rich_tracebacks=True),
    ],
)
logger = logging.getLogger(__name__)


class SalarySlipExtractor:
    """
    A class to extract salary slip data from PDF files and process it into a structured format.
    """

    def __init__(self, pdf_folder: str):
        """
        Initialize the extractor with the path to the folder containing PDF files.

        :param pdf_folder: Path to the directory containing salary slip PDFs.
        """
        self.pdf_folder = pdf_folder
        self.data_rows = []
        logger.info("Initialized SalarySlipExtractor for folder: %s", pdf_folder)

    def clean_value(self, value: str) -> str:
        """
        Clean extracted value by removing extra whitespace, line breaks, and pipes.

        :param value: Raw extracted value
        :return: Cleaned value
        """
        if not value:
            return None
        cleaned = re.sub(r"<br\s*/?>", " ", value)
        cleaned = cleaned.replace("|", "").strip()
        cleaned = re.sub(r"\s+", " ", cleaned)
        return cleaned if cleaned and cleaned != "-" else None

    def extract_from_table_row(self, markdown_text: str, field_name: str) -> str:
        """
        Extract value from markdown table format.

        :param markdown_text: Markdown text to search
        :param field_name: Field name to look for
        :return: Extracted value or None
        """
        pattern = rf"\|[*\s]*{re.escape(field_name)}[*\s]*\|([^|]+)\|"
        match = re.search(pattern, markdown_text, re.IGNORECASE)
        return self.clean_value(match.group(1)) if match else None

    def extract_numeric_value(self, markdown_text: str, field_name: str) -> str:
        """
        Extract numeric value (handles commas and decimals).

        :param markdown_text: Markdown text to search
        :param field_name: Field name to look for
        :return: Extracted numeric value or None
        """
        # Pattern to match field names with optional bold markers (**) and extract numeric values
        # Handles both regular and bolded field names, with optional <br> tags
        pattern = (
            rf"\*{{0,2}}{re.escape(field_name)}\*{{0,2}}\s*\|?\s*([\d,]+(?:\.\d{{2}})?)"
        )
        match = re.search(pattern, markdown_text, re.IGNORECASE)

        if match and match.group(1).strip():
            return match.group(1).strip()

        # Fallback: Try to find it in a table cell format (|**Field**|value|)
        table_pattern = rf"\|\s*\*{{0,2}}{re.escape(field_name)}\*{{0,2}}\s*\|\s*([\d,]+(?:\.\d{{2}})?)"
        table_match = re.search(table_pattern, markdown_text, re.IGNORECASE)

        return (
            table_match.group(1).strip()
            if table_match and table_match.group(1).strip()
            else None
        )

    def extract_data_from_pdf(self, pdf_path: str) -> Dict[str, str]:
        """
        Extract relevant data fields from a single PDF salary slip using pymupdf4llm.

        :param pdf_path: Full path to the PDF file.
        :return: Dictionary containing extracted data fields.
        """
        logger.info("Processing PDF: %s", pdf_path)
        data = {
            "File": os.path.basename(pdf_path),
            "Month": None,
            "Employee Name": None,
            "CNIC": None,
            "Designation": None,
            "Employment Basis": None,
            "Basic Salary": None,
            "House Rent Allowance": None,
            "Medical Allowance": None,
            "Utilities Allowance": None,
            "Bonus": None,
            "Fuel Reimbursements": None,
            "Gross Pay": None,
            "Income Tax": None,
            "Total Deduction": None,
            "Net Salary": None,
        }

        try:
            markdown_text = to_markdown(pdf_path)
            if not markdown_text:
                logger.warning("No text extracted from PDF: %s", pdf_path)
                return data

            logger.debug("Extracted markdown length: %d characters", len(markdown_text))

            basic_fields = {
                "Month": "Month of Salary",
                "Employee Name": "Employee Name",
                "CNIC": "CNIC No.",
                "Designation": "Designation",
                "Employment Basis": "Employment Basis",
            }

            for key, field_name in basic_fields.items():
                value = self.extract_from_table_row(markdown_text, field_name)
                if value:
                    data[key] = value
                    logger.info("✓ Extracted %s: %s", key, value)
                else:
                    logger.warning("✗ Could not extract: %s", key)

            numeric_fields = {
                "Basic Salary": "Basic Salary",
                "House Rent Allowance": "House Rent Allowance",
                "Medical Allowance": "Medical Allowance",
                "Utilities Allowance": "Utilities Allowance",
                "Bonus": "Bonus",
                "Fuel Reimbursements": "Fuel Reimbursements",
                "Gross Pay": "Gross Pay",
                "Income Tax": "Income Tax",
                "Total Deduction": "Total Deduction",
                "Net Salary": "Net Salary",
            }

            for key, field_name in numeric_fields.items():
                value = self.extract_numeric_value(markdown_text, field_name)
                if value:
                    data[key] = value
                    logger.info("✓ Extracted %s: %s", key, value)
                else:
                    logger.warning("✗ Could not extract: %s", key)

            for key in ["Bonus", "Fuel Reimbursements"]:
                if data[key] == "-" or data[key] == "":
                    data[key] = "0"

        except (FileNotFoundError, PermissionError, ValueError) as e:
            logger.error("Error processing PDF %s: %s", pdf_path, str(e), exc_info=True)
        return data

    def process_all_pdfs(self) -> None:
        """
        Process all PDF files in the specified folder and collect data.
        """
        logger.info("=" * 60)
        logger.info("Starting to process PDFs in folder: %s", self.pdf_folder)
        logger.info("=" * 60)

        try:
            pdf_files = [
                f for f in os.listdir(self.pdf_folder) if f.lower().endswith(".pdf")
            ]
        except (FileNotFoundError, PermissionError) as e:
            logger.error("Error accessing folder %s: %s", self.pdf_folder, str(e))
            return

        pdf_count = len(pdf_files)

        if pdf_count == 0:
            logger.warning("No PDF files found in folder: %s", self.pdf_folder)
            return

        logger.info("Found %d PDF files to process", pdf_count)

        for idx, file in enumerate(pdf_files, 1):
            logger.info("\n[%d/%d] Processing: %s", idx, pdf_count, file)
            logger.info("-" * 60)
            pdf_path = os.path.join(self.pdf_folder, file)
            data = self.extract_data_from_pdf(pdf_path)
            self.data_rows.append(data)

        logger.info("\n" + "=" * 60)
        logger.info("Completed processing %d PDF files", pdf_count)
        logger.info("=" * 60)

    def get_dataframe(self) -> pd.DataFrame:
        """
        Convert collected data rows into a Pandas DataFrame.

        :return: DataFrame with extracted data.
        """
        if not self.data_rows:
            logger.error("No data extracted. Process PDFs first.")
            raise ValueError("No data extracted. Process PDFs first.")
        logger.info("Converting %d data rows to DataFrame", len(self.data_rows))
        return pd.DataFrame(self.data_rows)


def add_totals_to_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Add a total row to the DataFrame for numeric columns.

    :param df: Input DataFrame.
    :return: DataFrame with an additional total row.
    """
    logger.info("Adding totals to DataFrame with %d rows", len(df))
    numeric_cols = [
        "Basic Salary",
        "House Rent Allowance",
        "Medical Allowance",
        "Utilities Allowance",
        "Bonus",
        "Fuel Reimbursements",
        "Gross Pay",
        "Income Tax",
        "Total Deduction",
        "Net Salary",
    ]

    try:
        for col in numeric_cols:
            if col in df.columns:
                df[col] = df[col].astype(str).str.replace(",", "").str.replace("-", "0")
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
                logger.debug("Converted column %s to numeric", col)

        totals = {col: df[col].sum() for col in numeric_cols if col in df.columns}
        totals["File"] = "TOTAL"
        totals["Month"] = ""
        totals["Employee Name"] = ""
        totals["CNIC"] = ""
        totals["Designation"] = ""
        totals["Employment Basis"] = ""

        total_row = pd.DataFrame([totals])
        result_df = pd.concat([df, total_row], ignore_index=True)
        logger.info("✅ Added total row to DataFrame")
        return result_df

    except (ValueError, TypeError) as e:
        logger.error("Error adding totals to DataFrame: %s", str(e))
        raise


def save_to_pretty_excel(df: pd.DataFrame, output_path: str) -> None:
    """
    Save the DataFrame to an Excel file with formatting, filters, and a table style.

    :param df: DataFrame to save.
    :param output_path: Path to save the Excel file.
    """
    logger.info("Saving DataFrame to Excel: %s", output_path)
    try:
        writer = pd.ExcelWriter(output_path, engine="openpyxl")
        df.to_excel(writer, index=False, sheet_name="Salary Slips")

        workbook: Workbook = writer.book
        worksheet = workbook.active

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        total_font = Font(bold=True, italic=True, size=11)
        total_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
        alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = thin_border

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except (TypeError, ValueError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column].width = adjusted_width

        for row in worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=worksheet.max_column,
        ):
            for cell in row:
                cell.border = thin_border
                if cell.column > 6:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        total_row_num = worksheet.max_row
        for cell in worksheet[total_row_num]:
            cell.font = total_font
            cell.fill = total_fill
            cell.border = thin_border

        tab = Table(
            displayName="SalaryTable", ref=f"A1:{worksheet.dimensions.split(':')[1]}"
        )
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        worksheet.add_table(tab)

        writer.close()
        logger.info("=" * 60)
        logger.info("✅ SUCCESS! Extracted data saved to: %s", output_path)
        logger.info("=" * 60)

    except (FileNotFoundError, PermissionError, ValueError) as e:
        logger.error("Error saving to Excel %s: %s", output_path, str(e))
        raise


def main(
    pdf_folder: str = "salary_slips", output_excel: str = "salary_slips_flattened.xlsx"
) -> None:
    """
    Main function to extract data from PDFs and save to a formatted Excel file.

    :param pdf_folder: Path to the folder with PDFs.
    :param output_excel: Path to save the Excel file (default: salary_slips_flattened.xlsx).
    """
    logger.info("\n" + "=" * 60)
    logger.info("SALARY SLIP EXTRACTOR - STARTED")
    logger.info("=" * 60)
    logger.info("Input folder: %s", pdf_folder)

    # Ensure output directory exists
    output_dir = "output"
    try:
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            logger.info("Created output directory: %s", output_dir)
    except (FileNotFoundError, PermissionError) as e:
        logger.error("Error creating output directory %s: %s", output_dir, str(e))
        raise

    output_path = os.path.join(output_dir, output_excel)
    logger.info("Output file: %s", output_path)
    logger.info("Using pymupdf4llm for PDF to Markdown conversion")

    try:
        extractor = SalarySlipExtractor(pdf_folder)
        extractor.process_all_pdfs()
        df = extractor.get_dataframe()

        logger.info("\n" + "=" * 60)
        logger.info("EXTRACTION SUMMARY")
        logger.info("=" * 60)
        logger.info("Total records extracted: %d", len(df))

        df_with_totals = add_totals_to_dataframe(df)
        save_to_pretty_excel(df_with_totals, output_path)

        logger.info("\n✅ PROCESS COMPLETED SUCCESSFULLY!")
        logger.info("=" * 60)

    except (ValueError, FileNotFoundError, PermissionError) as e:
        logger.error("❌ PROCESS FAILED: %s", str(e), exc_info=True)
        raise


if __name__ == "__main__":
    main("salary_slips")
